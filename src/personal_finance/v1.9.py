import pandas as pd
import os
from io import StringIO
import json
import openpyxl

# Define the directory containing the CSV files
directory = '/home/eikov/fin/OG-CSV'

# Define the pattern for file names
file_pattern = '{}{}.csv'  # {} will be replaced by the file number and +/- for the file type

# Initialize empty dictionaries to store dataframes for each month
spending_dfs = {}
income_dfs = {}

# Define the filters
filter1 = "KOVALIŪNAS EINARAS"
filter2 = "Einaras Kovaliūnas"  # Add your second filter value here

# Function to extract first 4 words from a string
def extract_first_4_words(text):
    if pd.isna(text):  # Check if the value is NaN
        return ""  # Return an empty string if NaN
    words = text.split()[:4]
    return ' '.join(words)


# Iterate over the file numbers from 1 to 12
for file_num in range(1, 13):
    month = str(file_num)
    spending_dfs[month] = pd.DataFrame(columns=["DATA", "SUMA", "MOKĖTOJO ARBA GAVĖJO PAVADINIMAS", "MOKĖJIMO PASKIRTIS"])
    income_dfs[month] = pd.DataFrame(columns=["DATA", "SUMA", "MOKĖTOJO ARBA GAVĖJO PAVADINIMAS", "MOKĖJIMO PASKIRTIS"])
    for file_type in ['+', '-']:
        filename = file_pattern.format(file_num, file_type)
        filepath = os.path.join(directory, filename)
        if os.path.exists(filepath):
            print(f"Processing file: {filename}")
            # Initialize an empty list to store the valid lines of the CSV file
            valid_lines = []
            skipped_rows = []

            # Read the CSV file line by line
            with open(filepath, 'r', encoding='utf-8') as file:
                for i, line in enumerate(file):
                    if i < 2:  # Skip the first two lines (first row and header row)
                        continue
                    try:
                        # Attempt to parse the line as CSV
                        pd.read_csv(StringIO(line), header=None, sep=';')  # Use header=None and specify separator as ';'
                        valid_lines.append(line)  # If successful, add the line to the list of valid lines
                    except pd.errors.ParserError:
                        print(f"Skipped line {i+1} in file {filename} due to ParserError: {line.strip()}")
                        skipped_rows.append((i+1, line.strip()))  # Store the line number and the content of the skipped row

            # Read the valid lines into a DataFrame, skipping the header line
            bank_data = None
            try:
                bank_data = pd.read_csv(StringIO(''.join(valid_lines)), header=None, sep=';')
            except pd.errors.ParserError:
                print(f"Encountered an error while parsing file {filename}. Skipping problematic rows.")
                if not valid_lines:
                    print("No valid data read due to errors.")
                    continue

            # If data was successfully read, extract specific columns and apply filter
            if bank_data is not None:
                bank_data = bank_data[[1, 3, 4, 9]]  # Extract only desired columns

                # Replace empty values with NaN
                bank_data.replace('-', pd.NA, inplace=True)

                # Rename columns
                bank_data.columns = ["DATA", "SUMA", "MOKĖTOJO ARBA GAVĖJO PAVADINIMAS", "MOKĖJIMO PASKIRTIS"]

                # Apply filters
                bank_data = bank_data[(bank_data["MOKĖTOJO ARBA GAVĖJO PAVADINIMAS"] != filter1) & 
                                      (bank_data["MOKĖTOJO ARBA GAVĖJO PAVADINIMAS"] != filter2)]

                # Filter out rows that are the same as column headers or contain all empty values
                bank_data = bank_data[~bank_data.apply(lambda row: all(row == ["DATA", "SUMA", "MOKĖTOJO ARBA GAVĖJO PAVADINIMAS", "MOKĖJIMO PASKIRTIS"]), axis=1)]

                # Extract first 4 words from specified columns
                bank_data["MOKĖTOJO ARBA GAVĖJO PAVADINIMAS"] = bank_data["MOKĖTOJO ARBA GAVĖJO PAVADINIMAS"].apply(extract_first_4_words)
                bank_data["MOKĖJIMO PASKIRTIS"] = bank_data["MOKĖJIMO PASKIRTIS"].apply(extract_first_4_words)

                # Append the DataFrame to the appropriate dictionary based on file type
                if file_type == '+':
                    income_dfs[month] = bank_data
                elif file_type == '-':
                    spending_dfs[month] = bank_data
            else:
                print(f"No valid data read from file {filename}.")

            # Display the DataFrame
            if file_type == '+':
                print(f"\nIncome DataFrame for month {month}:")
                print(income_dfs[month])
            elif file_type == '-':
                print(f"\nSpending DataFrame for month {month}:")
                print(spending_dfs[month])

                # Specify the folder path
                output_folder_income = "/home/eikov/fin/income"
                output_folder_spending = "/home/eikov/fin/spendings"

                # Create the folder if it doesn't exist
                os.makedirs(output_folder_income, exist_ok=True)
                os.makedirs(output_folder_spending, exist_ok=True)

                # Save DataFrames to JSON files with UTF-8 encoding
                income_json_filename = os.path.join( output_folder_income, f"income_month_{month}.json")
                spending_json_filename = os.path.join(output_folder_spending, f"spending_month_{month}.json")

                # Save DataFrames to Excel files
                income_excel_filename = os.path.join( output_folder_income, f"income_month_{month}.xlsx")
                spending_excel_filename = os.path.join(output_folder_spending, f"spending_month_{month}.xlsx")

                # Drop rows with NaN values
                income_dfs[month].dropna(inplace=True)
                spending_dfs[month].dropna(inplace=True)

                # Convert DataFrames to JSON strings
                income_json_str = income_dfs[month].to_json(orient='records', lines=True, default_handler=str)
                spending_json_str = spending_dfs[month].to_json(orient='records', lines=True, default_handler=str)
                """
                # Write JSON strings to files with UTF-8 encoding
                with open(income_json_filename, 'w+', encoding='utf-8') as income_file:
                    for line in income_json_str.split('\n'):
                        if line.strip():  # Check if line is not empty
                            income_file.write(line.strip() + '\n')

                with open(spending_json_filename, 'w+', encoding='utf-8') as spending_file:
                    for line in spending_json_str.split('\n'):
                        if line.strip():  # Check if line is not empty
                            spending_file.write(line.strip() + '\n')
                """
                # Save DataFrames to Excel files
                income_dfs[month].to_excel(income_excel_filename, index=False)
                spending_dfs[month].to_excel(spending_excel_filename, index=False)

                print(f"DataFrames saved to JSON files: {income_json_filename}, {spending_json_filename}")
                print(f"DataFrames saved to Excel files: {income_excel_filename}, {spending_excel_filename}")

            # Display skipped rows
            if skipped_rows:
                print(f"\nSkipped rows in file {filename}:")
                for row_num, row_content in skipped_rows:
                    print(f"Row {row_num}: {row_content}")
        else:
            print(f"File {filename} not found. Skipping...")

print("All files processed.")



#Merge spending seperate xlsx files to one
directory = '/home/eikov/fin/spendings'
output_directory = '/home/eikov/fin/analysis'

# List all files in the directory
files = os.listdir(directory)

# Filter out only the Excel files (files ending with .xlsx)
excel_files = [file for file in files if file.endswith('.xlsx')]

# Create a new workbook
merged_workbook = openpyxl.Workbook()

# Loop through each file and add it as a new sheet to the workbook
for i, file_name in enumerate(excel_files):
    # Load the workbook from file
    workbook = openpyxl.load_workbook(os.path.join(directory, file_name))
    
    # Get the active sheet from the loaded workbook
    sheet = workbook.active
    
    # Copy the active sheet to the merged workbook
    if i == 0:
        # For the first sheet, use the default sheet name 'Sheet'
        merged_sheet = merged_workbook.active
    else:
        # For subsequent sheets, create a new sheet with the file name as the sheet name
        merged_sheet = merged_workbook.create_sheet(title=f"Sheet{i+1}")
    
    # Copy data from the original sheet to the merged sheet
    for row in sheet.iter_rows(values_only=True):
        merged_sheet.append(row)

# Save the merged workbook to the output directory
output_filename = os.path.join(output_directory, "merged_spending_data.xlsx")
merged_workbook.save(output_filename)


#Merge income seperate xlsx files to one
directory = '/home/eikov/fin/income'
output_directory = '/home/eikov/fin/analysis'

# List all files in the directory
files = os.listdir(directory)

# Filter out only the Excel files (files ending with .xlsx)
excel_files = [file for file in files if file.endswith('.xlsx')]

# Create a new workbook
merged_workbook = openpyxl.Workbook()

# Loop through each file and add it as a new sheet to the workbook
for i, file_name in enumerate(excel_files):
    # Load the workbook from file
    workbook = openpyxl.load_workbook(os.path.join(directory, file_name))
    
    # Get the active sheet from the loaded workbook
    sheet = workbook.active
    
    # Copy the active sheet to the merged workbook
    if i == 0:
        # For the first sheet, use the default sheet name 'Sheet'
        merged_sheet = merged_workbook.active
    else:
        # For subsequent sheets, create a new sheet with the file name as the sheet name
        merged_sheet = merged_workbook.create_sheet(title=f"Sheet{i+1}")
    
    # Copy data from the original sheet to the merged sheet
    for row in sheet.iter_rows(values_only=True):
        merged_sheet.append(row)

# Save the merged workbook to the output directory
output_filename = os.path.join(output_directory, "merged_income_data.xlsx")
merged_workbook.save(output_filename)